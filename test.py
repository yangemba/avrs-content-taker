from openpyxl import load_workbook


class TakeData(object):
    def __init__(self, file_name):
        wb = load_workbook(file_name)
        self.sheet = wb['Sheetname']

    def take_hat(self):
        max_columns = self.sheet.max_column
        # c = sheet.cell(row=1, column=1).value
        list_hat = [self.sheet.cell(row=1, column=i).value for i in range(1, max_columns + 1)]
        list_hat.append('comment')
        list_hat.append('links')
        return list_hat

    def take_other_rows(self):
        max_columns = self.sheet.max_column
        max_rows = self.sheet.max_row
        list_rows = []

        for row in range(2, max_rows+1):
            single_row = [self.sheet.cell(row=row, column=i).value for i in range(1, max_columns + 1)]
            list_rows.append(single_row)
        return list_rows
            # for i in range(1, max_columns + 1):
            #     list_rows.append(self.sheet.cell(row=row, column=i).value)
            #     return self.sheet.cell(row=row, column=i).value


if __name__ == "__main__":
    take_data = TakeData('test-flats.xlsx')
    print(take_data.take_hat())
    print('\n\n')
    print(take_data.take_other_rows())




