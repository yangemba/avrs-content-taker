from openpyxl import load_workbook
import requests
import asyncio
import concurrent.futures
import requests
import time
import bs4
import logging
from test import TakeData
import csv

start = time.time()


def take_value(sheet, row, column):
    return sheet.cell(row=row, column=column).value


def take_ides():
    wb = load_workbook('test-flats.xlsx')
    sheet = wb['Sheetname']
    max_row_index = sheet.max_row
    id_adv_list = [take_value(sheet, x, 1) for x in range(2, max_row_index + 1)]
    return id_adv_list


def take_images(response):
    soup = bs4.BeautifulSoup(response.text, "html.parser")
    for link in soup.find_all("a", {"style": "display:none;"}):
        href = link.get('href')
        logging.warning(f"HREF - {href}")
        if href:
            return href
        return 'no image'


def parse_info(response):
    first = response.text.split('id="descripRealty">')
    comment = first[1].split('</div>')[0].replace('\n', '').replace('            ', '').replace('          ', '')\
        .replace('\r1', '').replace('\r2', '').replace('\r3', '').replace('\r', '').replace('\r4', '')\
        .replace('<p>', '').replace('<b>', '').replace('\rB', '')
    image_list = take_images(response)
    return comment, image_list


id_list = take_ides()
response_list = []


def id_from_response(response):
    url = response.url
    try:
        first = str(url).split('flat/')[1]
        main_id = first.split('.ht')[0]
    except Exception as e:
        logging.warning(f"RESPONSE ID ERROR: {e}")
        return None
    return main_id


def req_request(url):
    try:
        response = requests.get(url=url, allow_redirects=False)
    except Exception as e:
        return None
    return response


async def scrap():
    with concurrent.futures.ThreadPoolExecutor(max_workers=100) as executor:
        loop = asyncio.get_event_loop()
        futures = [
            loop.run_in_executor(executor, req_request, f'http://avers.in.ua/flat/{i}.htm')
            for i in id_list
        ]
        for response in await asyncio.gather(*futures):
            if response.status_code != 200:
                response_list.append({f"{id_from_response(response)}": None})
                continue
            try:
                logging.warning(response.status_code)
                info = parse_info(response)
            except Exception as e:
                logging.warning(f"error: {e}")
                response_list.append({f"{id_from_response(response)}": 'exception'})
            if info:
                response_list.append({f"{id_from_response(response)}": info})


if __name__ == "__main__":
    loop = asyncio.get_event_loop()
    loop.run_until_complete(scrap())
    finish = time.time()
    #
    # for i in response_list:
    #     print(i.values())
    # #
    # print(len(response_list))
    # print(str(finish - start))

    with open('result.txt', 'w') as file:
        pass

    for i in response_list:
        with open('result.txt', 'a') as file:
            file.write(f'{i}\n')

    with open("result.xls") as file:
        pass

    take_data = TakeData('test-flats.xlsx')
    hat = take_data.take_hat()
    rows_list = take_data.take_other_rows()
    # print(rows_list)

    final_list = []

    for row in rows_list:
        for el in response_list:
            # print(f'###{el.get(row[0])}')
            if el.get(str(row[0])):
                final_list.append(row + list((el.get(str(row[0])))))
    # for i in final_list:
    #     print(f'{i}\n')

    hat_string = f'{"  ".join(hat)}\n'

    with open("final.csv", "w", ) as file:
        file.write(hat_string)

    for element in final_list:
        current_str_row = f'{"  ".join(map(str, element))}\n'
        with open("final.csv", "a", ) as file:
            file.write(current_str_row)




    # with open("final.csv", "a", newline="") as file:
    #     writer = csv.writer(file)
    #     writer.writerows(final_list)

# Todo try to understand why lenth of the list is null
