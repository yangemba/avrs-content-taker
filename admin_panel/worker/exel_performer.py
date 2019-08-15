from openpyxl import load_workbook
import asyncio
import concurrent.futures
import requests
import time
import bs4
import logging
from admin_panel.worker.test import TakeData

start = time.time()


class WorkExel(object):

    def __init__(self, path):
        self.path_to_file = path

    def take_value(self, sheet, row, column):
        return sheet.cell(row=row, column=column).value

    def take_ides(self):
        wb = load_workbook(self.path_to_file)
        sheet = wb['Sheetname']
        max_row_index = sheet.max_row
        id_adv_list = [self.take_value(sheet, x, 1) for x in range(2, max_row_index + 1)]
        return id_adv_list

    def take_images(self, response):
        soup = bs4.BeautifulSoup(response.text, "html.parser")
        for link in soup.find_all("a", {"style": "display:none;"}):
            href = link.get('href')
            logging.warning(f"HREF - {href}")
            if href:
                return href
            return 'no image'

    def parse_info(self, response):
        first = response.text.split('id="descripRealty">')
        comment = first[1].split('</div>')[0].replace('\n', '').replace('            ', '').replace('          ', '')\
            .replace('\r1', '').replace('\r2', '').replace('\r3', '').replace('\r', '').replace('\r4', '')\
            .replace('<p>', '').replace('<b>', '').replace('\rB', '')
        image_list = self.take_images(response)
        return comment, image_list

    id_list = take_ides()
    response_list = []

    def id_from_response(self, response):
        url = response.url
        try:
            first = str(url).split('flat/')[1]
            main_id = first.split('.ht')[0]
        except Exception as e:
            logging.warning(f"RESPONSE ID ERROR: {e}")
            return None
        return main_id

    def req_request(self, url):
        try:
            response = requests.get(url=url, allow_redirects=False)
        except Exception as e:
            return None
        return response

    async def scrap(self):
        with concurrent.futures.ThreadPoolExecutor(max_workers=100) as executor:
            loop = asyncio.get_event_loop()
            futures = [
                loop.run_in_executor(executor, self.req_request, f'http://avers.in.ua/flat/{i}.htm')
                for i in self.id_list
            ]
            for response in await asyncio.gather(*futures):
                if response.status_code != 200:
                    self.response_list.append({f"{self.id_from_response(response)}": None})
                    continue
                try:
                    logging.warning(response.status_code)
                    info = self.parse_info(response)
                except Exception as e:
                    logging.warning(f"error: {e}")
                    self.response_list.append({f"{self.id_from_response(response)}": 'exception'})
                if info:
                    self.response_list.append({f"{self.id_from_response(response)}": info})

    def main_perform(self):
        loop = asyncio.get_event_loop()
        loop.run_until_complete(self.scrap())
        take_data = TakeData(self.path_to_file)
        hat = take_data.take_hat()
        rows_list = take_data.take_other_rows()
        # print(rows_list)

        final_list = []

        for row in rows_list:
            for el in self.response_list:
        # print(f'###{el.get(row[0])}')
                if el.get(str(row[0])):
                    final_list.append(row + list((el.get(str(row[0])))))

        hat_string = f'{"  ".join(hat)}\n'

        # with open("result/final.csv", "w", ) as file:
        #     file.write(hat_string)
        #
        # for element in final_list:
        #     current_str_row = f'{"  ".join(map(str, element))}\n'
        #     with open("result/final.csv", "a", ) as file:
        #         file.write(current_str_row)

#
# if __name__ == "__main__":
#     loop = asyncio.get_event_loop()
#     # loop.run_until_complete(scrap())
#     finish = time.time()
#     #
#     # for i in response_list:
#     #     print(i.values())
#     # #
#     # print(len(response_list))
#     # print(str(finish - start))
#
#     with open('result.txt', 'w') as file:
#         pass
#
#     # for i in response_list:
#         with open('result.txt', 'a') as file:
#             file.write(f'{i}\n')
#
#     with open("result.xls") as file:
#         pass
#
#     take_data = TakeData('test-flats.xlsx')
#     hat = take_data.take_hat()
#     rows_list = take_data.take_other_rows()
#     # print(rows_list)
#
#     final_list = []
#
#     # for row in rows_list:
#         # for el in response_list:
#             # print(f'###{el.get(row[0])}')
#             # if el.get(str(row[0])):
#             #     final_list.append(row + list((el.get(str(row[0])))))
#     # for i in final_list:
#     #     print(f'{i}\n')
#
#     hat_string = f'{"  ".join(hat)}\n'
#
#     with open("final.csv", "w", ) as file:
#         file.write(hat_string)
#
#     for element in final_list:
#         current_str_row = f'{"  ".join(map(str, element))}\n'
#         with open("final.csv", "a", ) as file:
#             file.write(current_str_row)




    # with open("final.csv", "a", newline="") as file:
    #     writer = csv.writer(file)
    #     writer.writerows(final_list)

# Todo try to understand why lenth of the list is null
